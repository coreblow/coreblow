#!/usr/bin/env python3
"""
scraper/export.py
Data export engine — CSV, JSON, Excel (.xlsx)
"""

import csv
import json
import io
from datetime import datetime
from utils.logger import logger

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False


class DataExporter:
    """Export scraped data to CSV, JSON, or Excel formats."""

    @staticmethod
    def to_csv(data, fields=None):
        """Export data to CSV string.

        Args:
            data: list of dicts
            fields: optional list of column names

        Returns:
            CSV string
        """
        if not data:
            return ""

        if not fields:
            fields = list(data[0].keys())

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=fields, extrasaction="ignore")
        writer.writeheader()

        for row in data:
            flat = {}
            for k, v in row.items():
                if isinstance(v, (dict, list)):
                    flat[k] = json.dumps(v, ensure_ascii=False)
                else:
                    flat[k] = v
            writer.writerow(flat)

        result = output.getvalue()
        logger.info(f"📤 Exported {len(data)} rows to CSV ({len(result)} bytes)")
        return result

    @staticmethod
    def to_json(data, pretty=True):
        """Export data to JSON string.

        Args:
            data: list of dicts
            pretty: indent output

        Returns:
            JSON string
        """
        result = json.dumps(data, indent=2 if pretty else None, ensure_ascii=False, default=str)
        logger.info(f"📤 Exported {len(data)} records to JSON ({len(result)} bytes)")
        return result

    @staticmethod
    def to_excel(data, fields=None, sheet_name="Scraped Data"):
        """Export data to Excel (.xlsx) bytes.

        Args:
            data: list of dicts
            fields: optional list of column names
            sheet_name: worksheet name

        Returns:
            bytes (xlsx file content)
        """
        if not HAS_OPENPYXL:
            raise ImportError("openpyxl is required for Excel export. Install: pip install openpyxl")

        if not data:
            return b""

        if not fields:
            fields = list(data[0].keys())

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name

        # Header row (bold)
        for col, field in enumerate(fields, 1):
            cell = ws.cell(row=1, column=col, value=field)
            cell.font = openpyxl.styles.Font(bold=True)

        # Data rows
        for row_idx, row in enumerate(data, 2):
            for col_idx, field in enumerate(fields, 1):
                value = row.get(field, "")
                if isinstance(value, (dict, list)):
                    value = json.dumps(value, ensure_ascii=False)
                ws.cell(row=row_idx, column=col_idx, value=value)

        # Auto-width columns
        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter
            for cell in col:
                try:
                    cell_len = len(str(cell.value or ""))
                    max_len = max(max_len, min(cell_len, 50))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = max_len + 2

        output = io.BytesIO()
        wb.save(output)
        result = output.getvalue()
        logger.info(f"📤 Exported {len(data)} rows to Excel ({len(result)} bytes)")
        return result

    @staticmethod
    def save_to_file(content, filepath, mode="w"):
        """Save export content to a file.

        Args:
            content: string or bytes
            filepath: output file path
            mode: 'w' for text, 'wb' for binary
        """
        with open(filepath, mode) as f:
            f.write(content)
        logger.info(f"💾 Saved to {filepath}")

    @staticmethod
    def generate_filename(format_type, target_name=None):
        """Generate a timestamped export filename."""
        ts = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        prefix = target_name.replace(" ", "_").lower() if target_name else "export"
        ext = {"csv": "csv", "json": "json", "excel": "xlsx"}.get(format_type, format_type)
        return f"scraper_{prefix}_{ts}.{ext}"
